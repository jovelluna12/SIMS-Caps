# from logging import root
import datetime
from datetime import date, datetime 
from tkinter import simpledialog

import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import tkinter
from tracemalloc import start
import Employee
import Manager

import datetime
from datetime import datetime
import calendar

from tkinter import filedialog as fd
from tkcalendar import DateEntry
from PIL import Image, ImageTk
import Owner
import Product
import forecast
import randomNumGen
import pandas as pd
from openpyxl.workbook import Workbook
import os

import openpyxl


class InvortoryGUI:

    def __init__(self):
        self.InvorVal = None
        global PageOpen
        PageOpen = 1
        global PageOpen_Sub
        PageOpen_Sub=1
        global filter_from, filter_to,filter,batch,FilterList

    def search(self):
        result = self.Entry_Search.get()

    # Chick List and stack Start~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def Filter_close(self):
        self.InvorVal.grab_release()
        FilterList.destroy()

    def filter_results(self):
        global filter_from, filter_to,filter,batch,FilterList
        if str(filter_from.cget("state"))=="normal":
            from_filter=filter_from.get_date()
        else: 
            from_filter=None

        if str(filter_to.cget("state"))=="normal":
                to_filter=filter_to.get_date()
        else: 
            to_filter=None

            fill=filter.get()
            batch_code=batch.get()
            man=Manager.Manager()

        if fill=="In Transit":
            res=man.get_inTransit(from_filter,to_filter,batch_code)
            count = 0
            self.frame_Table.delete(*self.frame_Table.get_children())
            for x in res:
                count += 1
                self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
           

        if fill=="On Hand":
                res=man.get_OnHand(from_filter,to_filter,batch_code)
                count = 0
                self.frame_Table.delete(*self.frame_Table.get_children())
                for x in res:
                    count += 1
                    self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
                

        if fill=="None":
                res=man.listNone(from_filter,to_filter,batch_code)
                count = 0
                self.frame_Table.delete(*self.frame_Table.get_children())
                for x in res:
                    count += 1
                    self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
                

        self.Filter_close()

    def filter_GUI(self):
        global filter_from, filter_to,filter,batch,FilterList
        FilterList=Toplevel(self.InvorVal)
        FilterList.title("Filter out Results")
        FilterList.geometry("480x340")
        FilterList.resizable(False,False)
        FilterList.protocol("WM_DELETE_WINDOW",self.Filter_close)

        FilterList.grab_set()
        Label(FilterList,text="LIST FILTER",font=("Arial", 35, "bold")).place(x=10,y=10)
        Label(FilterList,text="Use the following settings to Filter out the Results.\nYou can clear this settings Later on.",font=("Arial", 12, "bold")).place(x=33,y=80)
        Label(FilterList,text="Select Filter:").place(x=33,y=150)
        filter = ttk.Combobox(FilterList,width=10,state='readonly')
        filter.place(x=100, y=150)
        filter.set("None")
        filter["values"]=("None","In Transit","On Hand")

        Label(FilterList,text="Select Batch:").place(x=277,y=150)
        batch = ttk.Combobox(FilterList,width=10,state='readonly')
        batch.place(x=350, y=150)

        prod=Product.product()
        res=prod.get_batch_Codes()
        batch_list=[x[0] for x in res]
        batch_list.insert(0,"None")
        batch.set("None")
        batch["values"]=(batch_list)

        def no_sel():
            filter_from.config(state='disabled')
            filter_to.config(state='disabled')

        def sel():
            filter_from.config(state='normal')
            filter_to.config(state='normal')

        var=IntVar()
        var.set(0)
        Label(FilterList,text="Configure Date Parameters").place(x=33,y=180)
        Radiobutton(FilterList,text="No Parameters",variable=var,value=0,command=no_sel).place(x=100,y=200)
        Radiobutton(FilterList,text="Allow Date",variable=var,value=1,command=sel).place(x=230,y=200)

        Label(FilterList,text="Recorded From :").place(x=33,y=230)
        filter_from = DateEntry(FilterList,state='disabled') 
        filter_from.place(x=140, y=230)

        Label(FilterList,text="Recorded To :").place(x=33,y=260)
        filter_to = DateEntry(FilterList,state='disabled') 
        filter_to.place(x=140, y=260)
            
        Button(FilterList,text="Filter",bg="green",command=self.filter_results).place(x=340,y=300)
        Button(FilterList,text="Cancel",command=self.Filter_close).place(x=390,y=300)

    def Click_List1(self):
        self.button_List.config(state="disabled")
        self.button_Stack.config(state="normal")
        self.button_Delivery.config(state="normal")
        self.button_Employee.config(state="normal")

        self.Frame_stack.pack_forget()
        self.Frame_Del.pack_forget()
        self.Frame_Empl.pack_forget()

        self.Frame_List.pack()
        self.Label_title = Label(self.Frame_List, text="List Page", font=("Arial", 15)).place(x=0, y=0)

        btn_filter_clear=Button(self.Frame_List,text="Filters",bg="green",command=self.filter_GUI)
        btn_filter_clear.place(x=890, y=0)

        def clear_filter():
            m1 = Manager.Manager()
            result = m1.inventoryList()
            self.frame_Table.delete(*self.frame_Table.get_children())
            count = 0
            for x in result:
                count += 1
                self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)

        btn_filter_clear=Button(self.Frame_List,text="Clear Filters",command=lambda: clear_filter())
        btn_filter_clear.place(x=940, y=0)

        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview")
        self.frame_Table = ttk.Treeview(self.Frame_List, height=23)
        self.frame_Table['columns'] = ("ID", "Name", "Status", "Price", "Quantity")
        self.frame_Table.column("#0", width=0, stretch=NO)
        self.frame_Table.column("ID", anchor=W, width=100, stretch=NO)
        self.frame_Table.column("Name", anchor=W, width=500, stretch=NO)
        self.frame_Table.column("Status", anchor=W, width=150, stretch=NO)
        self.frame_Table.column("Price", anchor=CENTER, width=131, stretch=NO)
        self.frame_Table.column("Quantity", anchor=E, width=150, stretch=NO)
        # Table Head
        self.frame_Table.heading("#0")
        self.frame_Table.heading("ID", text="ID", anchor=W)
        self.frame_Table.heading("Name", text="Product Name", anchor=W)
        self.frame_Table.heading("Status", text="Status", anchor=W)
        self.frame_Table.heading("Price", text="Price", anchor=CENTER)
        self.frame_Table.heading("Quantity", text="Quantity", anchor=W)
        self.frame_Table.place(x=0, y=30)

        m1 = Manager.Manager()
        result = m1.inventoryList()
        count = 0
        for x in result:
            count += 1
            self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
        

    def Click_Stack(self):
        self.button_List.config(state="normal")
        self.button_Stack.config(state="disabled")
        self.button_Delivery.config(state="normal")
        self.button_Employee.config(state="normal")

        self.Frame_List.pack_forget()
        self.Frame_Del.pack_forget()
        self.Frame_Empl.pack_forget()

        self.Frame_stack.pack(fill='both')
        self.Label_title = Label(self.Frame_stack, text="SALES Page", font=("Arial", 15)).place(x=0, y=0)
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview")
        self.frame_Table = ttk.Treeview(self.Frame_stack, height=23)
        self.frame_Table['columns'] = ("ID", "Name", "Items Sold", "Price", "Sales", "Date Purchased")
        self.frame_Table.column("#0", width=0, stretch=NO)
        self.frame_Table.column("ID", anchor=W, width=100, stretch=NO)
        self.frame_Table.column("Name", anchor=W, width=450, stretch=NO)
        self.frame_Table.column("Items Sold", anchor=CENTER, width=100, stretch=NO)
        self.frame_Table.column("Price", anchor=CENTER, width=131, stretch=NO)
        self.frame_Table.column("Sales", anchor=E, width=100, stretch=NO)
        self.frame_Table.column("Date Purchased", anchor=E, width=150, stretch=NO)
        # Table Head
        self.frame_Table.heading("#0")
        self.frame_Table.heading("ID", text="Purchase ID", anchor=W)
        self.frame_Table.heading("Name", text="Product Name", anchor=W)
        self.frame_Table.heading("Items Sold", text="Items Sold", anchor=CENTER)
        self.frame_Table.heading("Price", text="Price", anchor=CENTER)
        self.frame_Table.heading("Sales", text="Sales", anchor=W)
        self.frame_Table.heading("Date Purchased", text="Date Purchased", anchor=W)
        self.frame_Table.place(x=0, y=30)
        m1 = Manager.Manager()
        result = m1.productSales()
        count = 0
        for x in result:
            count += 1
            self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
    
    def refresh(self):
        self.InvorGUI.update()
        self.InvorGUI.after(100, self.refresh)


    def Add_Delivery1_close(self):
            global PageOpen
            self.Add_Delivery1.wm_attributes("-topmost", 0)
            if messagebox.askokcancel('Close', 'Are you sure you want to close this Page?\n Unsaved data will not be Saved'):
                self.InvorVal.grab_release()
                PageOpen=1
                self.Add_Delivery1.destroy()
                self.refresh
            else:
                self.Add_Delivery1.wm_attributes("-topmost", 1)

    def ClickDelivery_onClick(self):
        global PageOpen
        if PageOpen < 2:
            if len(self.frame_Table.get_children()) == 0:
                messagebox.showinfo("Error","Sorry there's no Item to Receive!")
            else:
                if self.frame_Table.focus()!='':
                    item = self.frame_Table.selection()[0]
                    batch = self.frame_Table.item(item)['values'][0]
                    prod = Product.product()

                    if "idd_list" not in locals() and "batch_list" not in locals() and "ref_list" not in locals() and "QtyDif_list" not in locals() and "remark_list" not in locals() and "item_id_list" not in locals():
                        idd_list=[]
                        batch_list=[]
                        ref_list=[]
                        QtyDif_list=[]
                        remark_list=[]

                        item_id_list=[]

                    batch = (batch,)
                    result = prod.retrieveBatch(batch)

                    self.Add_Delivery1= Toplevel(self.InvorVal)
                    self.Add_Delivery1.title("Confirm Delivery")
                    self.Add_Delivery1.geometry("800x550")
                    self.Add_Delivery1.resizable(False, False)
                    self.Add_Delivery1.protocol("WM_DELETE_WINDOW",self.Add_Delivery1_close)
                    self.Add_Delivery1.wm_attributes("-topmost", 1)
                    self.Add_Delivery1.grab_set()

                    self.Frame_Add = Frame(self.Add_Delivery1, width=800, height=200)
                    self.Frame_Add.place(x=0, y=0)

                    self.Frame_ListD = Frame(self.Add_Delivery1, width=800, height=320, highlightbackground="black",
                                            highlightthickness=1, padx=10, pady=10)
                    self.Frame_ListD.place(x=0, y=200)

                    global idd, namee, qty, price
                    idd = StringVar()
                    namee = StringVar()
                    qty = StringVar()
                    price = StringVar()

                    self.Product_ID_LA = Label(self.Frame_Add, text="Product ID")
                    self.Product_ID_EN = Entry(self.Frame_Add, width=10, textvariable=idd, borderwidth=4, state='disabled')
                    self.Product_ID_LA.place(x=40, y=70)
                    self.Product_ID_EN.place(x=40, y=90)

                    self.Product_Price_LA = Label(self.Frame_Add, text="Product Name")
                    self.Product_Price_EN = Entry(self.Frame_Add, width=45, textvariable=namee, borderwidth=4, state='disabled')
                    self.Product_Price_LA.place(x=115, y=70)
                    self.Product_Price_EN.place(x=115, y=90)

                    self.Product_Stack_LA=Label(self.Frame_Add,text="Price")
                    self.Product_price_EN= Entry(self.Frame_Add,width=10,textvariable=price,borderwidth=4,state='disabled')
                    self.Product_Stack_LA.place(x=400,y=70)
                    self.Product_price_EN.place(x=400,y=90)

                    self.Product_Stack_LA=Label(self.Frame_Add,text="Quantity")
                    self.Product_Stack_EN= Entry(self.Frame_Add,width=10,textvariable=qty,borderwidth=4,state='disabled')
                    self.Product_Stack_LA.place(x=475,y=70)
                    self.Product_Stack_EN.place(x=475,y=90)

                    self.Product_date_LA = Label(self.Frame_Add, text="Expiry Date")
                    self.Product_date_EN = DateEntry(self.Frame_Add, selectmode='day', width=20, state='disabled')
                    self.Product_date_LA.place(x=550, y=70)
                    self.Product_date_EN.place(x=550, y=90)

                    Label(self.Frame_Add, text="Confirming Delivery Received Product",font=("Arial", 30)).place(x=10, y=10)

                    self.frame_Table = ttk.Treeview(self.Frame_ListD, height=15)
                    self.frame_Table['columns'] = ("ID", "Name", "Price", "Quantity", "Order Date", "Expiration Date")
                    self.frame_Table.column("#0", width=0, stretch=NO)
                    self.frame_Table.column("ID", anchor=W, width=50)
                    self.frame_Table.column("Name", anchor=W, width=246)
                    self.frame_Table.column("Price", anchor=CENTER, width=100)
                    self.frame_Table.column("Quantity", anchor=E, width=80)
                    self.frame_Table.column("Order Date", anchor=E, width=150)
                    self.frame_Table.column("Expiration Date", anchor=E, width=150)

                    self.frame_Table.heading("#0")
                    self.frame_Table.heading("ID", text="ID", anchor=W)
                    self.frame_Table.heading("Name", text="Product Name", anchor=W)
                    self.frame_Table.heading("Price", text="Price", anchor=CENTER)
                    self.frame_Table.heading("Quantity", text="Quantity", anchor=W)
                    self.frame_Table.heading("Order Date", text="Order Date", anchor=W)
                    self.frame_Table.heading("Expiration Date", text="Expiration Date", anchor=W)

                    self.frame_Table.pack(fill='both')
                    self.frame_Table.grid(row=1, column=0)
                    self.frame_Table.configure(selectmode="browse")

                    def confirm_delivery():
                        prod = Product.product()
                        id = idd.get()
                        name = namee.get()
                        qtyy = qty.get()

                        return_goods=list(zip(idd_list,batch_list,ref_list,QtyDif_list,remark_list))
                        prod.return_to_sender(return_goods)

                        for item in self.frame_Table.get_children():
                            id = self.frame_Table.item(item)['values'][0]
                            name = self.frame_Table.item(item)['values'][1]
                            priceeee = self.frame_Table.item(item)['values'][2]
                            qtyyy = self.frame_Table.item(item)['values'][3]
                            datee = self.frame_Table.item(item)['values'][5]
                            prod.editDelivery(id, name, priceeee, qtyyy, datee)

                        self.frame_Table.delete(*self.frame_Table.get_children())
                        idd_list.clear()
                        batch_list.clear()
                        ref_list.clear()
                        QtyDif_list.clear()
                        remark_list.clear()
                        item_id_list.clear()

                        self.Add_Delivery1.destroy()

                    
                    self.button = Button(self.Frame_Add, text="Confirm Delivery", command=confirm_delivery)
                    self.button.place(x=600, y=150)

                    self.delivery_delete = Button(self.Frame_Add, text="Delete", command=lambda: delete_delivery())
                    self.delivery_delete.place(x=550, y=150)

                    def delete_delivery():
                        if len(self.frame_Table.selection())!=0:
                            selected_items=self.frame_Table.item(self.frame_Table.selection())
                            id=selected_items['values'][0]
                            a=Product.product()
                            a.delete_del(id)
                            self.frame_Table.delete(self.frame_Table.selection())
                        else:
                            a=Product.product()
                            id=batch
                            a.delete_del_batch(id)
                            self.frame_Table.delete(*self.frame_Table.get_children())
                    

                    def selectItem(event):
                        selected_item = self.frame_Table.selection()[0]
                        if selected_item in item_id_list:
                            self.Add_Delivery1.wm_attributes("-topmost", 0)
                            messagebox.showerror("Already Saved","Item Already Saved")
                            self.Add_Delivery1.wm_attributes("-topmost", 1)
                        
                        else:
                            selected_item = self.frame_Table.selection()[0]
                            id = self.frame_Table.item(selected_item)['values'][0]
                            name = self.frame_Table.item(selected_item)['values'][1]
                            pricee = self.frame_Table.item(selected_item)['values'][2]
                            quantity = self.frame_Table.item(selected_item)['values'][3]
                            expire = self.frame_Table.item(selected_item)['values'][5]

                            self.Product_ID_EN.config(state='normal')
                            self.Product_Price_EN.config(state='normal')
                            self.Product_Stack_EN.config(state='normal')
                            self.Product_date_EN.config(state='normal')
                            self.Product_price_EN.config(state='normal')

                            self.Product_ID_EN.delete(0,END)
                            self.Product_Price_EN.delete(0,END)
                            self.Product_Stack_EN.delete(0,END)
                            self.Product_price_EN.delete(0,END)

                            global New_QTY_lbl
                            New_QTY=Label(self.Frame_Add,text="QTY to Receive")
                            New_QTY_lbl= Entry(self.Frame_Add,width=10,textvariable=qty,borderwidth=4)
                            New_QTY.place(x=450,y=110)
                            New_QTY_lbl.place(x=450,y=130)

                            date = datetime.strptime(expire, '%Y-%m-%d')

                            self.Product_date_EN.set_date(date)
                            idd.set(id)
                            namee.set(name)
                            qty.set(quantity)
                            price.set(pricee)

                            self.Product_ID_EN.insert(0,idd.get())
                            self.Product_Price_EN.insert(0,namee.get())
                            self.Product_Stack_EN.insert(0,qty.get())
                            self.Product_price_EN.insert(0,price.get())

                            self.Product_Price_EN.config(state='disabled')
                            self.Product_price_EN.config(state='disabled')
                            self.Product_ID_EN.config(state='disabled')
                            self.Product_Stack_EN.config(state='disabled')

                            self.button.config(state='normal', command=saveChanges)

                    def saveChanges(): 
                        try:
                            if int(New_QTY_lbl.get()) < int(qty.get()):
                                QTY_diff=int(qty.get())-int(New_QTY_lbl.get())
                                remark_dialog=simpledialog.askstring("Remarks", "New QTY is less than Order QTY.\nQTY not Included will be marked Return to Sender\nEnter Remark", parent=self.Frame_Add)
                                if remark_dialog:
                                    prod=Product.product()
                                    ref=prod.get_ref_id(namee.get())
                                    ref=ref[0]
                                            
                                    idd_list.append(idd.get())
                                    batch_list.append(batch[0])
                                    ref_list.append(ref)
                                    QtyDif_list.append(QTY_diff)
                                    remark_list.append(remark_dialog)

                                    newQTY=int(qty.get())-QTY_diff

                                    selectedItem = self.frame_Table.selection()[0]
                                    item_id_list.append(selectedItem)
                                    x = self.frame_Table.item(selectedItem)['values'][4]
                                    self.frame_Table.item(selectedItem,text="a", values=(
                                    self.Product_ID_EN.get(), self.Product_Price_EN.get(), self.Product_price_EN.get(), newQTY, x, self.Product_date_EN.get_date()))
                                    
                                    self.Product_ID_EN.config(state='normal')
                                    self.Product_Price_EN.config(state='normal')
                                    self.Product_Stack_EN.config(state='normal')
                                    self.Product_date_EN.config(state='normal')
                                    self.Product_price_EN.config(state='normal')

                                    self.Product_ID_EN.delete(0,END)
                                    self.Product_Price_EN.delete(0,END)
                                    self.Product_Stack_EN.delete(0,END)
                                    self.Product_price_EN.delete(0,END)

                                    self.Product_Price_EN.config(state='disabled')
                                    self.Product_price_EN.config(state='disabled')
                                    self.Product_ID_EN.config(state='disabled')
                                    self.Product_Stack_EN.config(state='disabled')

                            else:
                                selectedItem = self.frame_Table.selection()[0]
                                x = self.frame_Table.item(selectedItem)['values'][4]
                                self.frame_Table.item(selectedItem,text="a", values=(
                                self.Product_ID_EN.get(), self.Product_Price_EN.get(), self.Product_price_EN.get(), self.Product_Stack_EN.get(), x, self.Product_date_EN.get_date()))
                        except(ValueError):
                            self.Add_Delivery1.wm_attributes("-topmost", 0)
                            messagebox.showerror("Invalid Input","Please Enter a Valid Input")
                            self.Add_Delivery1.wm_attributes("-topmost", 1)
                            
                    self.button = Button(self.Frame_Add, text="Save", command=saveChanges)
                    self.button.place(x=710, y=150)

                    count = 0
                    for i in result:
                        self.frame_Table.insert(parent='', index='end', iid=count, text=i, values=i)
                        count += 1

                    self.frame_Table.bind("<Double-1>", selectItem)
                    PageOpen += 1
                else:
                    messagebox.showinfo("Error","No Item Selected")
        else:
            messagebox.showinfo("Error","The Window is already Open!")

    def Click_Delivery(self):
        self.button_List.config(state="normal")
        self.button_Stack.config(state="normal")
        self.button_Delivery.config(state="disabled")
        self.button_Employee.config(state="normal")

        self.Frame_List.pack_forget()
        self.Frame_stack.pack_forget()
        self.Frame_Empl.pack_forget()
        self.Frame_Del.pack()

        self.Label_title = Label(self.Frame_Del, text="Delivery Page", font=("Arial", 15)).place(x=0, y=0)
        self.Button_Receive=Button(self.Frame_Del,text="Receive",padx=5,pady=2,width=10,height=0,bg='#54FA9B',command=self.ClickDelivery_onClick)
        self.Button_Receive.place(x=945,y=0)
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview")
        self.frame_Table = ttk.Treeview(self.Frame_Del, height=23)
        self.frame_Table['columns'] = ("ID", "Name", "Detail", "Price", "Quantity", "Arrival")
        self.frame_Table.column("#0", width=0, stretch=NO)
        self.frame_Table.column("ID", anchor=W, width=100, stretch=NO)
        self.frame_Table.column("Name", anchor=W, width=450, stretch=NO)
        self.frame_Table.column("Detail", anchor=W, width=150, stretch=NO)
        self.frame_Table.column("Price", anchor=CENTER, width=91, stretch=NO)
        self.frame_Table.column("Quantity", anchor=E, width=100, stretch=OFF)
        self.frame_Table.column("Arrival", anchor=E, width=140, stretch=OFF)
        # Table Head
        self.frame_Table.heading("#0")
        self.frame_Table.heading("ID", text="Batch Code", anchor=W)
        self.frame_Table.heading("Name", text="Product Name", anchor=W)
        self.frame_Table.heading("Detail", text="Detail", anchor=W)
        self.frame_Table.heading("Price", text="Price", anchor=CENTER)
        self.frame_Table.heading("Quantity", text="Quantity", anchor=W)
        self.frame_Table.heading("Arrival", text="Arrival Day", anchor=W)
        self.frame_Table.place(x=0, y=30)

        m1 = Employee.Employee()
        result = m1.viewDeliveryList()
        count = 0
        for x in result:
            count += 1
            self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
    
    def deleteEmp(self,id):
        confirm_Delete=messagebox.askyesno("Confirm Delete","Are you sure you want to Delete this Employee?")
        if confirm_Delete:
            own=Owner.Owner()
            own.deleteEmp(id)
            messagebox.showinfo("Changes Saved","Employee Deleted Succesfully")
            self.Add_Employ.destroy()
            global PageOpen
            PageOpen=1
        else:
            messagebox.showinfo("Changes not Saved","Changes weren't Saved")

    def editEmp(self,id,name,username,password,role):
        confirm_Edit=messagebox.askyesno("Confirm Save","Are you sure you want to Save Edited Changes?")
        if confirm_Edit:
            own=Owner.Owner()
            own.EditEmp(id,name,username,password,role)
            messagebox.showinfo("Changes Saved","Changes Saved")
        else:
            messagebox.showinfo("Changes not Saved","Changes weren't Saved")

        
    def View_close(self):
                global PageOpen
                self.Add_Employ.wm_attributes("-topmost", 0)
                if messagebox.askokcancel('Close', 'Are you sure you want to close the Notification Page all the data will not be Save?'):
                    self.InvorVal.grab_release()
                    self.Add_Employ.destroy()
                    PageOpen=1
                else:
                    self.Add_Employ.wm_attributes("-topmost", 1)

    def View_onClick(self):
            global PageOpen
            if self.frame_Table.focus()=='':
                messagebox.showerror("No Item Selected","Please Select an Item First")
            else:
                items=self.frame_Table.focus()
                if PageOpen < 2:
                    self.Add_Employ= Toplevel(self.InvorVal)
                    self.Add_Employ.title("Employee Page")
                    self.Add_Employ.geometry("800x550")
                    self.Add_Employ.resizable(False, False)
                    self.Add_Employ.protocol("WM_DELETE_WINDOW",self.View_close)
                    self.Add_Employ.wm_attributes("-topmost", 1)
                    self.Add_Employ.grab_set()

                    self.Frame_Empl_VIEW = Frame(self.Add_Employ, width=790, height=195)
                    self.Frame_Empl_VIEW.place(x=0, y=0)

                    self.Frame_ListE = Frame(self.Add_Employ, width=790, height=300, highlightbackground="black",
                                        highlightthickness=1, padx=10, pady=10)
                    self.Frame_ListE.place(x=0, y=200)

                    emp_id = StringVar()
                    namee = StringVar()
                    username = StringVar()
                    passwd = StringVar()

                    self.Employ_ID_LA = Label(self.Frame_Empl_VIEW, text="Employ ID")
                    self.Employ_ID_EN = Entry(self.Frame_Empl_VIEW, width=10, textvariable=emp_id, borderwidth=4)
                    self.Employ_ID_LA.place(x=40, y=70)
                    self.Employ_ID_EN.place(x=40, y=90)

                    self.Employ_Name_LA = Label(self.Frame_Empl_VIEW, text="Name")
                    self.Employ_Name_EN = Entry(self.Frame_Empl_VIEW, width=45, textvariable=namee, borderwidth=4)
                    self.Employ_Name_LA.place(x=115, y=70)
                    self.Employ_Name_EN.place(x=115, y=90)

                    self.Employ_Uname_LA = Label(self.Frame_Empl_VIEW, text="Username")
                    self.Employ_Uname_EN = Entry(self.Frame_Empl_VIEW, width=40, textvariable=username, borderwidth=4)
                    self.Employ_Uname_LA.place(x=40, y=120)
                    self.Employ_Uname_EN.place(x=40, y=140)

                    self.Empoly_Pass_LA=Label(self.Frame_Empl_VIEW,text="Password")
                    self.Empoly_Pass_EN= Entry(self.Frame_Empl_VIEW,width=40,textvariable=passwd,borderwidth=4)
                    self.Empoly_Pass_LA.place(x=300,y=120)
                    self.Empoly_Pass_EN.place(x=300,y=140)

                    self.Employee_Role_LA = Label(self.Frame_Empl_VIEW, text="Role:")
                    self.chosen_val_Edit = tk.StringVar(self.Frame_Empl_VIEW)
                    self.Role_emplo = ttk.Combobox(self.Frame_Empl_VIEW, textvariable=self.chosen_val_Edit)
                    self.Role_emplo['values'] = ('Cashier', 'Manager')
                    self.Role_emplo.place(x=420,y=90)
                    self.Employee_Role_LA.place(x=420,y=70)

                    def edit():
                        self.Employ_Name_EN.config(state='normal')
                        self.Employ_Uname_EN.config(state='normal')
                        self.Empoly_Pass_EN.config(state='normal')
                        self.Role_emplo.config(state='readonly')

                    def save():
                        self.Add_Employ.wm_attributes("-topmost", 0)
                        id=self.Employ_ID_EN.get()
                        name=self.Employ_Name_EN.get()
                        uname=self.Employ_Uname_EN.get()
                        password=self.Empoly_Pass_EN.get()
                        role=self.Role_emplo.get()
                        self.editEmp(int(id),name, uname, password, role)
                        self.Employ_Name_EN.config(state='disabled')
                        self.Employ_Uname_EN.config(state='disabled')
                        self.Empoly_Pass_EN.config(state='disabled')
                        self.Add_Employ.wm_attributes("-topmost", 1)

                    Button_Edit=Button(self.Frame_Empl_VIEW,text="Edit",padx=5,pady=2,width=10,height=0,bg='#54FA9B',command=lambda: edit())
                    Button_Edit.place(x=600,y=150)
                        
                    Button_Save=Button(self.Frame_Empl_VIEW,text="Save",padx=5,pady=2,width=10,height=0,bg='#54FA9B',command=lambda: save())
                    Button_Save.place(x=600,y=120)
                        
                    Button_Delete=Button(self.Frame_Empl_VIEW,text="Delete",padx=5,pady=2,width=10,height=0,bg='#54FA9B',command=lambda: self.deleteEmp(int(idd.get())))
                    Button_Delete.place(x=700,y=120)

                    Button_Cancel=Button(self.Frame_Empl_VIEW,text="Close",padx=5,pady=2,width=10,height=0,bg='#54FA9B',command=self.View_close)
                    Button_Cancel.place(x=700,y=150)

                    if user_role=='Manager':
                        Button_Edit.config(state='disabled')
                        Button_Save.config(state='disabled')
                        Button_Delete.config(state='disabled')

                    Label(self.Frame_Empl_VIEW, text="Employee",font=("Arial", 30)).place(x=10, y=10)

                    id=self.frame_Table.item(items)['values'][0]
                    emp=Manager.Manager()
                    res=emp.selectEmp(id)
                    val1,val2,val3,val4,val5=[tuple for tuple in res]
                    emp_id.set(val1)
                    self.Employ_ID_EN.insert(0, emp_id.get())
                    self.Employ_ID_EN.config(state='disabled')
                    namee.set(val2)
                    self.Employ_Name_EN.insert(0, namee.get())
                    self.Employ_Name_EN.config(state='disabled')
                    self.chosen_val_Edit.set(val5)
                    self.Role_emplo.config(state='disabled')
                    username.set(val3)
                    self.Employ_Uname_EN.insert(0,username.get())
                    self.Employ_Uname_EN.config(state='disabled')
                    passwd.set(val4)
                    self.Empoly_Pass_EN.insert(0,passwd.get())
                    self.Empoly_Pass_EN.config(state='disabled')

                    self.emp_Table = ttk.Treeview(self.Frame_ListE, height=15)
                    self.emp_Table['columns'] = ("Time In","Time Out","Date")
                    self.emp_Table.column("#0", width=0, stretch=NO)
                    self.emp_Table.column("Time In", anchor=W, width=246)
                    self.emp_Table.column("Time Out", anchor=W, width=280)
                    self.emp_Table.column("Date", anchor=W, width=230)

                    self.emp_Table.heading("#0")
                    self.emp_Table.heading("Time In", text="Time In", anchor=W)
                    self.emp_Table.heading("Time Out", text="Time Out", anchor=W)
                    self.emp_Table.heading("Date", text="Date", anchor=W)
                    scrollbar = ttk.Scrollbar(self.Frame_ListE, orient="vertical", command=self.emp_Table.yview)
                    self.emp_Table.configure(yscrollcommand=scrollbar.set)
                    self.emp_Table.pack(fill='both')
                    self.emp_Table.grid(row=1, column=0)
                    scrollbar.grid(row=1, column=1, sticky="ns")

                    EmpData=Employee.Employee()
                    ress=EmpData.getAttendance(val1)
                    count = 0
                    for x in ress:
                        self.emp_Table.insert(parent='', index='end', iid=count, text=(x[3],x[4],x[2]), values=(x[3],x[4],x[2]))
                        count += 1

                    PageOpen += 1
                else:
                    messagebox.showinfo("Error","The Window is already Open!")

    def Click_Employee(self):
        self.button_List.config(state="normal")
        self.button_Stack.config(state="normal")
        self.button_Delivery.config(state="normal")
        self.button_Employee.config(state="disabled")

        self.Frame_List.pack_forget()
        self.Frame_stack.pack_forget()
        self.Frame_Del.pack_forget()

        self.Frame_Empl.pack()
        self.Label_title = Label(self.Frame_Empl, text="Employee Page", font=("Arial", 15)).place(x=0, y=0)
        self.Button_Emplo=Button(self.Frame_Empl,text="Employee Page",padx=5,pady=2,width=10,height=0,bg='#54FA9B',command=self.View_onClick)
        self.Button_Emplo.place(x=945,y=0)

        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview")

        self.frame_Table = ttk.Treeview(self.Frame_Empl, height=23)
        self.frame_Table['columns'] = ("ID", "Name", "Username", "Detail")
        self.frame_Table.column("#0", width=0, stretch=NO)
        self.frame_Table.column("ID", anchor=W, width=98, stretch=NO)
        self.frame_Table.column("Name", anchor=W, width=420, stretch=NO)
        self.frame_Table.column("Username", anchor=W, width=280, stretch=NO)
        self.frame_Table.column("Detail", anchor=CENTER, width=230, stretch=NO)

        # Table Head
        self.frame_Table.heading("#0")
        self.frame_Table.heading("ID", text="Employee ID", anchor=W)
        self.frame_Table.heading("Name", text="Name", anchor=W)
        self.frame_Table.heading("Username", text="Username", anchor=W)
        self.frame_Table.heading("Detail", text="Role", anchor=CENTER)
        self.frame_Table.place(x=0, y=30)

        m1 = Manager.Manager()
        result = m1.viewEMPList()
        count = 0
        for x in result:
            count += 1
            self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
        
        if len(self.frame_Table.get_children())==0:
            self.Button_Emplo.config(state='disabled')
        else:
            self.Button_Emplo.config(state='normal')

        if len(self.frame_Table.get_children())==0:
            self.Button_Emplo.config(state='disabled')
        else:
            self.Button_Emplo.config(state='normal')


    def add_to_products(self):
        global order_date, arrival_date
        ProdName = self.Product_CODE_EN.get()
        date = self.Product_date_EN.get_date()
        expiry_date = self.Product_EXdate_EN.get_date()
        quantity = self.Product_Stack_EN.get()
        ProductIDD = randomNumGen.generateProductID()
        price = float(self.Product_Price_EN.get())
        order_date = self.Product_date_EN.get_date()
        arrival_date = self.Product_Arrive_EN.get_date()

        self.Product_Price_EN.config(state='normal')
        self.Product_Price_EN.delete(0,END)
        self.Product_Price_EN.config(state='disabled')


        var1 = "ProductID_list"
        var2 = "ProdName_list"
        var3 = "quantity_list"
        var4 = "price_list"
        var5 = "date_list"
        var6 = "expiry_date_list"
        var7 = "ref_id_list"

        if var1 not in globals() and var2 not in globals() and var3 not in globals() and var4 not in globals() and var5 not in globals() and var6 not in globals() and var7 not in globals():
            global ProductID_list, ProdName_list, ref_id_list, quantity_list, price_list, date_list, expiry_date_list, status_list, arrival_date_list, order_date_list

            ProductID_list = []
            ProdName_list = []
            quantity_list = []
            price_list = []
            date_list = []
            expiry_date_list = []
            order_date_list = []
            status_list = []
            ref_id_list = []

        prod = Product.product()
        ref_id = prod.get_ref_id(ProdName)

        ProductID_list.append(int(ProductIDD))
        ProdName_list.append(ProdName)
        quantity_list.append(int(quantity))
        price_list.append(price)
        date_list.append(date)
        expiry_date_list.append(expiry_date)
        order_date_list.append(order_date)
        status_list.append("In Transit")
        ref_id_list.append(ref_id[0])

        self.frame_Table.insert(parent='', index='end', iid=ProductIDD,
                                text=(ProductIDD, ProdName, price, quantity, date, expiry_date),
                                values=(ProductIDD, ProdName, price, quantity, date, expiry_date))

        self.Product_CODE_EN.delete(0, 'end')
        self.Product_Stack_EN.delete(0, 'end')
        self.Product_EXdate_EN.delete(0, 'end')
        self.Product_Price_EN.delete(0, 'end')

        self.AddDeliveries = Button(self.Frame_Add, text="Add to Delivery List", padx=15, pady=5,
                                    command=self.Add_Deliveries)
        self.AddDeliveries.place(x=544, y=150)

        self.button_Add.config(command=self.add_to_products)

    def Add_Deliveries(self):
        self.Add_Delivery.wm_attributes("-topmost", 0)
        if messagebox.askokcancel('Comfirm', 'Please make sure if the product to delivery and Quantity is correct'):
            if 'batch_code' not in locals():
                batch_code = randomNumGen.generateBatchCode()
            if 'batch_code_list' not in locals():
                batch_code_list = []
            for i in ProductID_list:
                batch_code_list.append(batch_code)

            item_tuple = list(
                zip(ProductID_list, ref_id_list, ProdName_list, quantity_list, price_list, status_list, batch_code_list,
                    expiry_date_list))

            values = (batch_code, order_date, arrival_date, 'In Transit')

            b = Product.product()
            b.add_deliveryBatch(values)

            b.addMany_Del(item_tuple)

            self.frame_Table.delete(*self.frame_Table.get_children())

            ProductID_list.clear()
            batch_code_list.clear()
            ProdName_list.clear()
            quantity_list.clear()
            price_list.clear()
            date_list.clear()
            expiry_date_list.clear()
            order_date_list.clear()
            status_list.clear()
            PagaOpen = 1 
            self.Add_Delivery.destroy()
        else:
            self.Add_Delivery.wm_attributes("-topmost", 1)

    def search_delivery(self,event):
        value=event.widget.get()
        if value !='':
            data=[]
            for item in range(len(lst)):
                if value.lower() in lst[item][1].lower():
                    data.append(lst[item][1])
               
            event.widget['values']=data

        else:
            a = Product.product()
            re = a.returnall()
            n = 1
            event.widget['values']=([x[n] for x in re])

  
    def Add_on_close(self):
            global PageOpen
            self.Add_Delivery.wm_attributes("-topmost", 0)
            if messagebox.askokcancel('Close', 'Are you sure you want to close Window?'):
                self.InvorVal.grab_release()
                self.Add_Del['bg']='#54FA9B'
                PageOpen=1
                self.Add_Delivery.destroy()
            else:
                self.Add_Employ.wm_attributes("-topmost", 1)

    def Click_Add_Delivery(self):
        global PageOpen
        if PageOpen < 2:
            self.Add_Del['bg']='gray'
            self.Add_Delivery = Toplevel(self.InvorVal)
            self.Add_Delivery.title("Add Products on Delivery")
            self.Add_Delivery.geometry("800x550")
            self.Add_Delivery.resizable(False, False)
            self.Add_Delivery.protocol("WM_DELETE_WINDOW", self.Add_on_close)
            self.Add_Delivery.wm_attributes("-topmost", 1)
            self.Add_Delivery.grab_set()

            global btn, frame
            btn = self.Add_Del
            frame = self.Add_Delivery


            global lst
            a = Product.product()
            lst = a.returnall()
            n = 1

            self.Frame_Add = Frame(self.Add_Delivery, width=800, height=200)
            self.Frame_Add.place(x=0, y=0)

            self.chosen_val = tk.StringVar(self.Frame_Add)
            self.chosen_val.set("Select Product")
            global price_entry
            price_entry = tk.StringVar()

            self.APD = Label(self.Frame_Add, text="Add Products on Delivery", font=("Arial", 40)).place(x=10, y=5)
            self.Product_CODE_LA = Label(self.Frame_Add, text="Select Product Name")
            self.Product_CODE_EN = ttk.Combobox(self.Frame_Add, textvariable=self.chosen_val, width=50)
            self.Product_CODE_LA.place(x=20, y=70)
            self.Product_CODE_EN.place(x=20, y=90)

            if lst != "empty":
                self.Product_CODE_EN['values'] = ([x[n] for x in lst])

            self.Product_CODE_EN.bind("<<ComboboxSelected>>", self.setPrice)    
            self.Product_CODE_EN.bind("<KeyRelease>",self.search_delivery)

            self.Product_date_LA = Label(self.Frame_Add, text="Date")
            self.Product_date_EN = DateEntry(self.Frame_Add, selectmode='day', width=20)
            self.Product_date_LA.place(x=20, y=120)
            self.Product_date_EN.place(x=20, y=140)

            self.Product_Exdate_LA = Label(self.Frame_Add, text="Expiration Date")
            self.Product_EXdate_EN = DateEntry(self.Frame_Add, selectmode='day', width=20)
            self.Product_Exdate_LA.place(x=200, y=120)
            self.Product_EXdate_EN.place(x=200, y=140)

            self.Product_Arrive_LA = Label(self.Frame_Add, text="Arrival Date")
            self.Product_Arrive_EN = DateEntry(self.Frame_Add, selectmode='day', width=20)
            self.Product_Arrive_LA.place(x=380, y=120)
            self.Product_Arrive_EN.place(x=380, y=140)

            
            self.Product_Price_LA = Label(self.Frame_Add, text="Price")
            self.Product_Price_EN = Entry(self.Frame_Add, width=20, borderwidth=4, textvariable=price_entry,
                                        state='disabled')
            self.Product_Price_LA.place(x=380, y=70)
            self.Product_Price_EN.place(x=380, y=90)

            self.Product_Stack_LA = Label(self.Frame_Add, text="Quantity")
            self.Product_Stack_EN = Entry(self.Frame_Add, width=20, borderwidth=4)
            self.Product_Stack_LA.place(x=520, y=70)
            self.Product_Stack_EN.place(x=520, y=90)

            self.button_Add = Button(self.Frame_Add, text="Add", padx=20, pady=5, command=self.add_to_products)
            self.button_Add.place(x=700, y=150)

            self.button_Delete = Button(self.Frame_Add, text="Delete", padx=20, pady=5, command=self.Delete)
            self.button_Delete.place(x=700, y=110)

            self.Frame_List1 = Frame(self.Add_Delivery, width=800, height=320, highlightbackground="black",
                                    highlightthickness=1, padx=10, pady=10)
            self.Frame_List1.place(x=0, y=200)
            # Table
            self.frame_Table = ttk.Treeview(self.Frame_List1, height=15)
            self.frame_Table['columns'] = ("ID", "Name", "Price", "Quantity", "Order Date", "Expiration Date")
            self.frame_Table.column("#0", width=0, stretch=NO)
            self.frame_Table.column("ID", anchor=W, width=50)
            self.frame_Table.column("Name", anchor=W, width=246)
            self.frame_Table.column("Price", anchor=CENTER, width=100)
            self.frame_Table.column("Quantity", anchor=E, width=80)
            self.frame_Table.column("Order Date", anchor=E, width=150)
            self.frame_Table.column("Expiration Date", anchor=E, width=140)
            # Table Head
            self.frame_Table.heading("#0")
            self.frame_Table.heading("ID", text="ID", anchor=W)
            self.frame_Table.heading("Name", text="Product Name", anchor=W)
            self.frame_Table.heading("Price", text="Price", anchor=CENTER)
            self.frame_Table.heading("Quantity", text="Quantity", anchor=W)
            self.frame_Table.heading("Order Date", text="Order Date", anchor=W)
            self.frame_Table.heading("Expiration Date", text="Expiration Date", anchor=W)
            scrollbar = ttk.Scrollbar(self.Frame_List1, orient="vertical", command=self.frame_Table.yview)
            self.frame_Table.configure(yscrollcommand=scrollbar.set)
            self.frame_Table.pack(fill='both')
            self.frame_Table.grid(row=1, column=0)
            scrollbar.grid(row=1, column=1, sticky="ns")

            PageOpen += 1
            self.Add_Delivery.mainloop()
        else:
            messagebox.showinfo("Error","The Window already Open!")
    
    def Employee_on_close(self):
        global PageOpen
        if messagebox.askokcancel('Close', 'Are you sure you want to close Window?'):
            PageOpen=2
            self.Add_Employee.destroy()

    # employee
    def Click_AddS_Em(self):
        username = self.Employee_Username_EN.get()
        password = self.Employee_Password_EN.get ()
        Fname = self.Employee_Lname_EN.get()
        role = self.chosen_val.get()

        man = Manager.Manager()
        id = randomNumGen.generateEmpID()
        man.AddEmp(id, Fname, username, password, role)
        self.Add_Employee.destroy()
        global PageOpen
        PageOpen=0
    
    def Employee_on_close(self):
            global PageOpen
            self.Add_Employee.wm_attributes("-topmost", 0)
            if messagebox.askokcancel('Close', 'Are you sure you want to close Window?'):
                self.InvorVal.grab_release()
                self.button_Add_Em['bg']='#54FA9B'
                PageOpen=1
                self.Add_Employee.destroy()
            else:
                self.Add_Employee.wm_attributes("-topmost", 1)

    def Click_Add_Em(self):
        global PageOpen
        if PageOpen<2:
            if user_role=='Manager': messagebox.showerror("Access not Granted","Only the Owner is allowed to Open this")
            else:
                self.button_Add_Em['bg']='gray'
                self.Add_Employee = Toplevel(self.InvorVal)
                self.Add_Employee.title("Employeee!")
                self.Add_Employee.geometry("500x400")
                self.Add_Employee.resizable(False, False)
                self.Add_Employee.protocol("WM_DELETE_WINDOW", self.Employee_on_close)
                self.Add_Employee.wm_attributes("-topmost", 1)
                self.Add_Employee.grab_set()

                global btn, frame
                btn = self.button_Add_Em
                frame = self.Add_Employee

                self.Frame_Add_Em = Frame(self.Add_Employee, width=800, height=500, )
                self.Frame_Add_Em.place(x=0, y=0)

                Frma = Label(self.Frame_Add_Em, text="Add Employee", width=20, font=("Arial", 35), anchor=W)
                Frma.place(x=20, y=20)

                self.Fname = StringVar()
                self.Employee_Lname_LA = Label(self.Frame_Add_Em, text="Full Name:")
                self.Employee_Lname_EN = Entry(self.Frame_Add_Em, width=60, borderwidth=4, textvariable=self.Fname)
                self.Employee_Lname_LA.place(x=60, y=110)
                self.Employee_Lname_EN.place(x=60, y=130)

                self.Employee_Username_LA = Label(self.Frame_Add_Em, text="Username:")
                self.username = StringVar()
                self.Employee_Username_EN = Entry(self.Frame_Add_Em, width=60, textvariable=self.username, borderwidth=4)
                self.Employee_Username_LA.place(x=60, y=160)
                self.Employee_Username_EN.place(x=60, y=180)

                self.Employee_Password_LA = Label(self.Frame_Add_Em, text="Password:")
                self.password = StringVar()
                self.Employee_Password_EN = Entry(self.Frame_Add_Em, width=60, textvariable=self.password, show="*",
                                                borderwidth=4)
                self.Employee_Password_LA.place(x=60, y=210)
                self.Employee_Password_EN.place(x=60, y=230)

                self.Employee_Role_LA = Label(self.Frame_Add_Em, text="Role:")
                self.chosen_val = tk.StringVar(self.Frame_Add_Em)
                self.chosen_val.set("Select Role")
                self.Role = ttk.Combobox(self.Frame_Add_Em, textvariable=self.chosen_val, state='readonly')
                self.Role['values'] = ('Cashier', 'Manager')
                self.Role.place(x=60, y=280)
                self.Employee_Role_LA.place(x=60, y=260)

                self.button_Add = Button(self.Frame_Add_Em, text="Add", padx=20, pady=5, command=self.Click_AddS_Em)
                self.button_Add.place(x=360, y=330)
                PageOpen += 1
        else:
            messagebox.showinfo("Error","The Window already Open!")


    def AddProduct(self):
        self.frame_Table.delete(*self.frame_Table.get_children())

        Prod = Product.product()
        Prod.addMany(vals)

    # stack
    def DoneAdd_Product(self):

        ProductName = self.Stack_Product_Name_EN.get()

        Quantity = self.Stack_Product_Size_EN.get()
        price = price_entry.get()

        id = randomNumGen.generateProductID()

        if Quantity.isdigit():
            Quantityy = int(Quantity)

        pricee = float(int(float(price)))
        global vals
        vals = (id, ProductName, pricee, Quantityy)

        global count

        id = []
        name = []
        quantity = []
        price = []
        batch_code = []
        status = []

        if 'count' not in globals():
            count = 0
        else:
            count += 1
        self.frame_Table.insert(parent='', index='end', iid=count, text=vals, values=(vals))
        vals = ()
        batch = randomNumGen.generateBatchCode()

        for child in self.frame_Table.get_children():
            val = self.frame_Table.item(child)["values"]
            idd = randomNumGen.generateProductID()

            id.append(idd)
            batch_code.append(batch)
            name.append(val[1])
            quantity.append(val[2])
            price.append(val[3])
            status.append("Added")

            vals = list(zip(id, name, quantity, price, batch_code, status))

        self.button_Finish.config(state='normal', command=self.AddProduct)

    def setPrice(self, event):
        choice = event.widget.get()
        res = [idx for idx, x in enumerate(lst) if x[1] == choice]
        item = lst[res[0]]
        price = item[2]
        self.Product_Price_EN.config(state='normal')
        self.Product_Price_EN.delete(0,END)
        self.Product_Price_EN.insert(0,price)
        self.Product_Price_EN.config(state='disabled')

    def Add_Stack_on_close(self):
        global PageOpen
        self.Add_Stack.wm_attributes("-topmost", 0)
        if messagebox.askokcancel('Close', 'Are you sure you want to close Window?'):
            self.InvorVal.grab_release()
            self.button_Add_Pm['bg']='#54FA9B'
            PageOpen=1
            self.Add_Stack.destroy()
        else:
            self.Add_Stack.wm_attributes("-topmost", 1)

    def Click_Add_Product(self):
        global PageOpen
        if PageOpen<2:
            self.button_Add_Pm['bg']='gray'
            self.Add_Stack = Toplevel(self.InvorVal)
            self.Add_Stack.title("Add Product")
            self.Add_Stack.geometry("800x540")
            self.Add_Stack.resizable(False, False)
            self.Add_Stack.wm_attributes("-topmost", 1)
            self.Add_Stack.grab_set()

            global btn, frame
            btn = self.button_Add_Pm
            frame = self.Add_Stack

            self.Add_Stack.protocol("WM_DELETE_WINDOW", self.Add_Stack_on_close)

            self.Frame_Add_St = Frame(self.Add_Stack, width=800, height=505, )
            self.Frame_Add_St.place(x=0, y=0)

            self.Frma = Label(self.Frame_Add_St, text="Add Product", width=20, font=("Arial", 35), anchor=W)
            self.Frma.place(x=10, y=10)

            a = Product.product()

            self.chosen_val = tk.StringVar(self.Frame_Add_St)
            self.chosen_val.set("Select Product")

            global lst
            lst = a.returnall()
            n = 1

            self.Stack_Product_Name_LA = Label(self.Frame_Add_St, text="Product Name:")
            self.Stack_Product_Name_EN = ttk.Combobox(self.Frame_Add_St, textvariable=self.chosen_val)
            self.Stack_Product_Name_LA.place(x=20, y=80)
            self.Stack_Product_Name_EN.place(x=20, y=100)
            self.Stack_Product_Name_EN.config(width=50)

            if lst != 'empty':
                self.Stack_Product_Name_EN['values'] = ([x[n] for x in lst])

            self.Frame_ListAP = Frame(self.Add_Stack, width=800, height=320, padx=10, pady=10, highlightbackground="black",
                                    highlightthickness=1)
            self.Frame_ListAP.place(x=0, y=150)
            # Table
            self.frame_Table = ttk.Treeview(self.Frame_ListAP, height=17)
            self.frame_Table['columns'] = ("ID", "Name", "Price")
            self.frame_Table.column("#0", width=0, stretch=NO)
            self.frame_Table.column("ID", anchor=W, width=90)
            self.frame_Table.column("Name", anchor=W, width=555)
            self.frame_Table.column("Price", anchor=CENTER, width=120)
            # Table Head
            self.frame_Table.heading("#0")
            self.frame_Table.heading("ID", text="ID", anchor=W)
            self.frame_Table.heading("Name", text="Product Name", anchor=W)
            self.frame_Table.heading("Price", text="Price", anchor=W)
            scrollbar = ttk.Scrollbar(self.Frame_ListAP, orient="vertical", command=self.frame_Table.yview)
            self.frame_Table.configure(yscrollcommand=scrollbar.set)
            self.frame_Table.pack(fill='both')
            self.frame_Table.grid(row=1, column=0)
            scrollbar.grid(row=1, column=1, sticky="ns")

            a = Product.product()
            res = a.returnall()

            for i in res:
                self.frame_Table.insert(parent='', index='end', iid=i[0], text=i, values=i)

            def AddProduct_ChangeName(event):
                global val
                val = self.chosen_val.get()
                val = (val)

            def search():
                a = Product.product()
                res = a.return_one(val)

                self.frame_Table.delete(*self.frame_Table.get_children())
                for i in res:
                    self.frame_Table.insert(parent='', index='end', iid=i[0], text=i, values=i)

            def edit():
                var = self.Stack_Product_Name_EN.get()
                self.Click_Edit_Ref(var)

            self.Stack_Product_Name_EN.bind("<<ComboboxSelected>>", AddProduct_ChangeName)
            self.Stack_Product_Name_EN.bind('<KeyRelease>',self.search_delivery)

            self.button_Find = Button(self.Frame_Add_St, text="Search", padx=20, pady=5, command=search)
            self.button_Find.place(x=350, y=90)

            self.button_Add = Button(self.Frame_Add_St, text="Add Product", padx=20, pady=5, command=self.Click_Add_Ref)
            self.button_Add.place(x=500, y=90)

            self.button_Edit = Button(self.Frame_Add_St, text="Edit", padx=20, pady=5, command=edit)
            self.button_Edit.place(x=623, y=90)

            self.button_Delete = Button(self.Frame_Add_St, text="Delete", padx=20, pady=5, command=self.Delete)
            self.button_Delete.place(x=700, y=90)

            self.Add_Stack.update()
            PageOpen += 1

        else:
            messagebox.showinfo("Error","The Window already Open!")
    
    #Sale>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    def Sale_Filter_close(self):
        self.Add_Notify.grab_release()
        self.Sale_FilterList.wm_attributes("-topmost", 1)
        self.Sale_FilterList.destroy()
        

    def Sale_Filter_results(self):
        # global filter_from, filter_to,filter,batch,self.Sale_FilterList
        if str(filter_from.cget("state"))=="normal":
            from_filter=filter_from.get_date()
        else: 
            from_filter=None

        if str(filter_to.cget("state"))=="normal":
                to_filter=filter_to.get_date()
        else: 
            to_filter=None

            fill=filter.get()
            batch_code=batch.get()
            man=Manager.Manager()

        if fill=="In Transit":
            res=man.get_inTransit(from_filter,to_filter,batch_code)
            count = 0
            self.frame_Table.delete(*self.frame_Table.get_children())
            for x in res:
                count += 1
                self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
           

        if fill=="On Hand":
                res=man.get_OnHand(from_filter,to_filter,batch_code)
                count = 0
                self.frame_Table.delete(*self.frame_Table.get_children())
                for x in res:
                    count += 1
                    self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
                

        if fill=="None":
                res=man.listNone(from_filter,to_filter,batch_code)
                count = 0
                self.frame_Table.delete(*self.frame_Table.get_children())
                for x in res:
                    count += 1
                    self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
                
        self.Sale_Filter_close()

    def Sale_Filter_GUI(self):
        self.InvorVal.grab_release()
        self.Add_Notify.wm_attributes("-topmost", 0)

        self.Sale_FilterList=Toplevel(self.Add_Notify)
        self.Sale_FilterList.title("Filter out Results")
        self.Sale_FilterList.geometry("480x340")
        self.Sale_FilterList.resizable(False,False)
        self.Sale_FilterList.protocol("WM_DELETE_WINDOW",self.Sale_Filter_close)
        self.Sale_FilterList.grab_set()
        self.Sale_FilterList.wm_attributes("-topmost", 1)
        
        Label(self.Sale_FilterList,text="SALE REPORTS FILTER",font=("Arial", 25, "bold")).place(x=10,y=10)
        Label(self.Sale_FilterList,text="Use the following settings to Filter out the Results.\nYou can clear this settings Later on.",font=("Arial", 12, "bold")).place(x=33,y=80)
        Label(self.Sale_FilterList,text="Select Filter:").place(x=33,y=150)
        filter = ttk.Combobox(self.Sale_FilterList,width=10,state='readonly')
        filter.place(x=100, y=150)
        filter.set("None")
        filter["values"]=("None","In Transit","On Hand")

        Label(self.Sale_FilterList,text="Select Batch:").place(x=277,y=150)
        batch = ttk.Combobox(self.Sale_FilterList,width=10,state='readonly')
        batch.place(x=350, y=150)

        prod=Product.product()
        res=prod.get_batch_Codes()
        batch_list=[x[0] for x in res]
        batch_list.insert(0,"None")
        batch.set("None")
        batch["values"]=(batch_list)

        def no_sel():
            filter_from.config(state='disabled')
            filter_to.config(state='disabled')

        def sel():
            filter_from.config(state='normal')
            filter_to.config(state='normal')

        var=IntVar()
        var.set(0)
        Label(self.Sale_FilterList,text="Configure Date Parameters").place(x=33,y=180)
        Radiobutton(self.Sale_FilterList,text="No Parameters",variable=var,value=0,command=no_sel).place(x=100,y=200)
        Radiobutton(self.Sale_FilterList,text="Allow Date",variable=var,value=1,command=sel).place(x=230,y=200)

        Label(self.Sale_FilterList,text="Recorded From :").place(x=33,y=230)
        filter_from = DateEntry(self.Sale_FilterList,state='disabled') 
        filter_from.place(x=140, y=230)

        Label(self.Sale_FilterList,text="Recorded To :").place(x=33,y=260)
        filter_to = DateEntry(self.Sale_FilterList,state='disabled') 
        filter_to.place(x=140, y=260)
            
        Button(self.Sale_FilterList,text="Filter",bg="green",command=self.Sale_Filter_results).place(x=340,y=300)
        Button(self.Sale_FilterList,text="Cancel",command=self.Sale_Filter_close).place(x=390,y=300)
    
    #Inventory>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    def Inven_Filter_close(self):
        self.Add_Notify.grab_release()
        self.Inven_FilterList.wm_attributes("-topmost", 1)
        self.Inven_FilterList.destroy()
        
    def Inven_Filter_results(self):
        # global filter_from, filter_to,filter,batch,self.Sale_FilterList
        if str(filter_from.cget("state"))=="normal":
            from_filter=filter_from.get_date()
        else: 
            from_filter=None

        if str(filter_to.cget("state"))=="normal":
                to_filter=filter_to.get_date()
        else: 
            to_filter=None

            fill=filter.get()
            batch_code=batch.get()
            man=Manager.Manager()

        if fill=="In Transit":
            res=man.get_inTransit(from_filter,to_filter,batch_code)
            count = 0
            self.frame_Table.delete(*self.frame_Table.get_children())
            for x in res:
                count += 1
                self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
           

        if fill=="On Hand":
                res=man.get_OnHand(from_filter,to_filter,batch_code)
                count = 0
                self.frame_Table.delete(*self.frame_Table.get_children())
                for x in res:
                    count += 1
                    self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
                

        if fill=="None":
                res=man.listNone(from_filter,to_filter,batch_code)
                count = 0
                self.frame_Table.delete(*self.frame_Table.get_children())
                for x in res:
                    count += 1
                    self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
                
        self.Inven_Filter_close()

    def Inven_Filter_GUI(self):
        self.InvorVal.grab_release()
        self.Add_Notify.wm_attributes("-topmost", 0)

        self.Inven_FilterList=Toplevel(self.Add_Notify)
        self.Inven_FilterList.title("Filter out Results")
        self.Inven_FilterList.geometry("480x340")
        self.Inven_FilterList.resizable(False,False)
        self.Inven_FilterList.protocol("WM_DELETE_WINDOW",self.Inven_Filter_close)
        self.Inven_FilterList.grab_set()
        self.Inven_FilterList.wm_attributes("-topmost", 1)
        
        Label(self.Inven_FilterList,text="Inventory FILTER",font=("Arial", 25, "bold")).place(x=10,y=10)
        Label(self.Inven_FilterList,text="Use the following settings to Filter out the Results.\nYou can clear this settings Later on.",font=("Arial", 12, "bold")).place(x=33,y=80)
        Label(self.Inven_FilterList,text="Select Filter:").place(x=33,y=150)
        filter = ttk.Combobox(self.Inven_FilterList,width=10,state='readonly')
        filter.place(x=100, y=150)
        filter.set("None")
        filter["values"]=("None","In Transit","On Hand")

        Label(self.Inven_FilterList,text="Select Batch:").place(x=277,y=150)
        batch = ttk.Combobox(self.Inven_FilterList,width=10,state='readonly')
        batch.place(x=350, y=150)

        prod=Product.product()
        res=prod.get_batch_Codes()
        batch_list=[x[0] for x in res]
        batch_list.insert(0,"None")
        batch.set("None")
        batch["values"]=(batch_list)

        def no_sel():
            filter_from.config(state='disabled')
            filter_to.config(state='disabled')

        def sel():
            filter_from.config(state='normal')
            filter_to.config(state='normal')

        var=IntVar()
        var.set(0)
        Label(self.Inven_FilterList,text="Configure Date Parameters").place(x=33,y=180)
        Radiobutton(self.Inven_FilterList,text="No Parameters",variable=var,value=0,command=no_sel).place(x=100,y=200)
        Radiobutton(self.Inven_FilterList,text="Allow Date",variable=var,value=1,command=sel).place(x=230,y=200)

        Label(self.Inven_FilterList,text="Recorded From :").place(x=33,y=230)
        filter_from = DateEntry(self.Inven_FilterList,state='disabled') 
        filter_from.place(x=140, y=230)

        Label(self.Inven_FilterList,text="Recorded To :").place(x=33,y=260)
        filter_to = DateEntry(self.Inven_FilterList,state='disabled') 
        filter_to.place(x=140, y=260)
            
        Button(self.Inven_FilterList,text="Filter",bg="green",command=self.filter_results).place(x=340,y=300)
        Button(self.Inven_FilterList,text="Cancel",command=self.Inven_Filter_close).place(x=390,y=300)

    #Delivery>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    def Del_Filter_close(self):
        self.Add_Notify.grab_release()
        self.Del_FilterList.wm_attributes("-topmost", 1)
        self.Del_FilterList.destroy()
        
    def Del_Filter_results(self):
        # global filter_from, filter_to,filter,batch,self.Sale_FilterList
        if str(filter_from.cget("state"))=="normal":
            from_filter=filter_from.get_date()
        else: 
            from_filter=None

        if str(filter_to.cget("state"))=="normal":
                to_filter=filter_to.get_date()
        else: 
            to_filter=None

            fill=filter.get()
            batch_code=batch.get()
            man=Manager.Manager()

        if fill=="In Transit":
            res=man.get_inTransit(from_filter,to_filter,batch_code)
            count = 0
            self.frame_Table.delete(*self.frame_Table.get_children())
            for x in res:
                count += 1
                self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
           

        if fill=="On Hand":
                res=man.get_OnHand(from_filter,to_filter,batch_code)
                count = 0
                self.frame_Table.delete(*self.frame_Table.get_children())
                for x in res:
                    count += 1
                    self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
                

        if fill=="None":
                res=man.listNone(from_filter,to_filter,batch_code)
                count = 0
                self.frame_Table.delete(*self.frame_Table.get_children())
                for x in res:
                    count += 1
                    self.frame_Table.insert(parent='', index='end', iid=count, text=x, values=x)
                
        self.Del_Filter_close()

    def Del_Filter_GUI(self):
        self.InvorVal.grab_release()
        self.Add_Notify.wm_attributes("-topmost", 0)

        self.Del_FilterList=Toplevel(self.Add_Notify)
        self.Del_FilterList.title("Filter out Results")
        self.Del_FilterList.geometry("480x340")
        self.Del_FilterList.resizable(False,False)
        self.Del_FilterList.protocol("WM_DELETE_WINDOW",self.Del_Filter_close)
        self.Del_FilterList.grab_set()
        self.Del_FilterList.wm_attributes("-topmost", 1)
        
        Label(self.Del_FilterList,text="Delivery FILTER",font=("Arial", 25, "bold")).place(x=10,y=10)
        Label(self.Del_FilterList,text="Use the following settings to Filter out the Results.\nYou can clear this settings Later on.",font=("Arial", 12, "bold")).place(x=33,y=80)
        Label(self.Del_FilterList,text="Select Filter:").place(x=33,y=150)
        filter = ttk.Combobox(self.Del_FilterList,width=10,state='readonly')
        filter.place(x=100, y=150)
        filter.set("None")
        filter["values"]=("None","In Transit","On Hand")

        Label(self.Del_FilterList,text="Select Batch:").place(x=277,y=150)
        batch = ttk.Combobox(self.Del_FilterList,width=10,state='readonly')
        batch.place(x=350, y=150)

        prod=Product.product()
        res=prod.get_batch_Codes()
        batch_list=[x[0] for x in res]
        batch_list.insert(0,"None")
        batch.set("None")
        batch["values"]=(batch_list)

        def no_sel():
            filter_from.config(state='disabled')
            filter_to.config(state='disabled')

        def sel():
            filter_from.config(state='normal')
            filter_to.config(state='normal')

        var=IntVar()
        var.set(0)
        Label(self.Del_FilterList,text="Configure Date Parameters").place(x=33,y=180)
        Radiobutton(self.Del_FilterList,text="No Parameters",variable=var,value=0,command=no_sel).place(x=100,y=200)
        Radiobutton(self.Del_FilterList,text="Allow Date",variable=var,value=1,command=sel).place(x=230,y=200)

        Label(self.Del_FilterList,text="Recorded From :").place(x=33,y=230)
        filter_from = DateEntry(self.Del_FilterList,state='disabled') 
        filter_from.place(x=140, y=230)

        Label(self.Del_FilterList,text="Recorded To :").place(x=33,y=260)
        filter_to = DateEntry(self.Del_FilterList,state='disabled') 
        filter_to.place(x=140, y=260)
            
        Button(self.Del_FilterList,text="Filter",bg="green",command=self.Del_Filter_results).place(x=340,y=300)
        Button(self.Del_FilterList,text="Cancel",command=self.Del_Filter_close).place(x=390,y=300)

    def Delete(self):
        select = self.frame_Table.selection()[0]
        id=self.frame_Table.item(select)["values"][0]
        prod=Product.product()
        prod.delete_ProdRef(id)
        self.frame_Table.delete(select)

    def Add_Notify_on_close(self):
        global PageOpen
        self.Add_Notify.wm_attributes("-topmost", 0)
        if messagebox.askokcancel('Close', 'Are you sure you want to close Window?'):
            self.InvorVal.grab_release()
            self.btn_Notification['bg']='#54FA9B'
            PageOpen=1
            self.Add_Notify.destroy()
        else:
            self.Add_Notify.wm_attributes("-topmost", 1)

    # start UI for Export---------------
    def notify_UI(self):
        global PageOpen, to,fromm, num_days
        to=None
        fromm=None

        now = datetime.now()
        num_days = calendar.monthrange(now.year, now.month)[1]

        if PageOpen<2:
            self.btn_Notification['bg']='gray'
            self.Add_Notify = Toplevel(self.InvorVal)
            global btn, frame
            btn = self.btn_Notification
            frame = self.Add_Notify
            self.Add_Notify.protocol("WM_DELETE_WINDOW", self.Add_Notify_on_close)
            self.Add_Notify.title("Export to Spreadsheet")
            self.Add_Notify.geometry("800x583")
            self.Add_Notify.resizable(False,False)
            self.Add_Notify.wm_attributes("-topmost", 1)
            self.Add_Notify.grab_set()

            self.Frame_Add_nofi = Frame(self.Add_Notify, width=800, height=200)
            self.Frame_Add_nofi.place(x=0, y=0)

            self.Frma = Label(self.Frame_Add_nofi, text="Export to Spreadsheet", width=20, font=("Arial", 35), anchor=W)
            self.Frma.place(x=10, y=10)
                
            selected = StringVar()
            Label(self.Add_Notify, text="Select What to Export").place(x=20, y=80)
            reports = ttk.Combobox(self.Add_Notify, width=20)
            reports.place(x=20, y=100)
            reports['values'] = ("Sales", "Inventory", "Delivery")
            
            # Label(self.Add_Notify, text="Click this Button to Start Exporting").place(x=550, y=100)
            export = Button(self.Add_Notify, text="Export", state='disabled')
            export.place(x=740, y=97)

            Table_BOX=Frame(self.Add_Notify,highlightbackground="black", highlightthickness=3)
            Table_BOX.place(x=0,y=140,relwidth=1.0,relheight=0.76)
            self.export_Table = ttk.Treeview(Table_BOX,height=25)
            self.export_Table['columns'] = ("Invoice Number","Purchase ID", "Item", "Quantity", "Total Price", "Discount", "Date Purchased")
            self.export_Table.column("#0", width=0, stretch=NO)
            self.export_Table.column("Invoice Number", anchor=W,)
            self.export_Table.column("Purchase ID", anchor=W,)
            self.export_Table.column("Item", anchor=W)
            self.export_Table.column("Quantity", anchor=E,)
            self.export_Table.column("Total Price", anchor=E,)
            self.export_Table.column("Discount", anchor=E,)
            self.export_Table.column("Date Purchased", anchor=E, )

            self.export_Table.heading("#0")
            self.export_Table.heading("Invoice Number", text="Invoice Number", anchor=W)
            self.export_Table.heading("Purchase ID", text="Purchase ID", anchor=W)
            self.export_Table.heading("Item", text="Item", anchor=W)
            self.export_Table.heading("Quantity", text="Quantity", anchor=W)
            self.export_Table.heading("Total Price", text="Total Price", anchor=W)
            self.export_Table.heading("Discount", text="Discount", anchor=W)
            self.export_Table.heading("Date Purchased", text="Date Purchased", anchor=W)
            scrollbar = ttk.Scrollbar(Table_BOX, orient="vertical", command=self.export_Table)
            scrollbar.pack(side="right", fill="y")

            self.export_Table.configure(yscrollcommand=scrollbar.set)    
            self.export_Table.pack()

            def export_report():
                self.Add_Notify.wm_attributes("-topmost", 0)
                report_type = reports.get()
                man = Manager.Manager()

                SALES_TEMPLATE = 'SALES_TEMPLATE.xlsx'
                INVENTORY_TEMPLATE = 'INVENTORY_TEMPLATE.xlsx'
                DELIVERY_TEMPLATE = 'PO_TEMPLATE.xlsx'

                if report_type == 'Sales':
                    print("this")
                    # from_month=scope_to.get()
                    # year=scope_year.get()
                    # if option=="Monthly":
                    #     month_num=datetime.strptime(scope.get(), '%B').month
                    #     date_obj=datetime(int(year), month_num, 1)

                    #     date_str_to = date_obj.strftime('%Y-%m-%d')

                    #     from_mnth=datetime.strptime(from_month, '%B').month
                    #     from_mnth_obj=datetime(int(year), from_mnth, calendar.monthrange(int(year), from_mnth)[1])

                    #     frm_mtnh_str=from_mnth_obj.strftime('%Y-%m-%d')
                    #     result = man.get_export_data(report_type,date_str_to,frm_mtnh_str)

                    #     df = pd.DataFrame(result, columns=['Invoice Number', 'ID', 'Item', 'Quantity','Unit Price','Discount', 'Date Purchased','Total Price'])
                    # if option=="Day":
                    #     month=from_month

                    #     month_num=datetime.strptime(month, '%B').month
                    #     date_obj=datetime(int(year), month_num, int(scope.get()))
                    #     date_str = date_obj.strftime('%Y-%m-%d')

                    #     result = man.get_export_data(report_type,date_str,date_str)
                    #     df = pd.DataFrame(result, columns=['Invoice Number', 'ID', 'Item', 'Quantity','Unit Price','Discount', 'Date Purchased','Total Price'])

                    # wb=openpyxl.load_workbook(SALES_TEMPLATE)
                    # ws=wb.worksheets[0]
                    # start_row=12
                    # for row in df.iterrows():
                    #     for column in range(len(row[1])):
                    #         ws.cell(row=start_row,column=column+2).value=row[1][column]
                    #     start_row+=1
                    #     ws.insert_rows(start_row,1)
                    # path=fd.asksaveasfilename(defaultextension=".xlsx")
                    # wb.save(path)

                elif report_type == 'Inventory':
                    result = man.get_export_data(report_type,None,None)
                    df = pd.DataFrame(result, columns=['Reference ID', "Item", "Price", "Remaining Quantity"])
                    wb=openpyxl.load_workbook(INVENTORY_TEMPLATE)
                    ws=wb.worksheets[0]
                    start_row=5 
                    for row in df.iterrows():
                        for column in range(len(row[1])):
                            ws.cell(row=start_row,column=column+2).value=row[1][column]
                        start_row+=1
                        ws.insert_rows(start_row,1)

                    path=fd.asksaveasfilename(defaultextension=".xlsx")
                    wb.save(path) 

                elif report_type == "Delivery":
                    result = man.get_export_data(report_type,None,None)
                    df = pd.DataFrame(result, columns=['Batch Code', 'Item', 'Quantity', 'Price', 'Status'])
                    wb=openpyxl.load_workbook(DELIVERY_TEMPLATE)    
                    ws=wb.worksheets[0]
                    start_row=28
                    for row in df.iterrows():
                        for column in range (len(row[1])):
                            ws.cell(row=start_row,column=column+2).value=row[1][column]
                        start_row+=1
                        ws.insert_rows(start_row,1)

                    path=fd.asksaveasfilename(defaultextension=".xlsx")
                    wb.save(path)        

                # if report_type == 'Forecast':
                #     result, value = man.get_export_data(report_type,None,None)
                #     df=pd.DataFrame(result,columns=['Id','Item','Quantity','Price'])
                #     df.insert(4,"30 Day Forecast",value)

                #     path=fd.asksaveasfilename(defaultextension=".xlsx")
                #     df.to_excel(path, str(report_type))
                    # title = str.lower(report_type) + str(date.today()) + '.xlsx'
                    # df.to_excel(title, str(report_type))
                    # message = "Saved to ", title
                messagebox.showinfo("Exported Successfully", "Saved to " + path)

            
            def reports_callback(event):
                export.config(state='normal',command=lambda:export_report())
                self.export_Table.delete(*self.export_Table.get_children())
                report_type = reports.get()

                man = Manager.Manager()
                result = man.get_export_data(report_type,None,None)
                global Filterbutton, Clear_Filterbutton,Inven_Filterbutton, Inven_Clear_Filterbutton,Deli_Filterbutton, Deli_Clear_Filterbutton 
                if report_type == 'Sales':
                    try:
                        Inven_Filterbutton.place_forget()
                        Inven_Clear_Filterbutton.place_forget()
                        Deli_Filterbutton.place_forget()
                        Deli_Clear_Filterbutton.place_forget()
                    except(NameError):
                        pass

                    
                    Filterbutton=Button(self.Add_Notify,text="Sales Filter",command=self.Sale_Filter_GUI)
                    Filterbutton.place(x=200, y=95)
                    Clear_Filterbutton=Button(self.Add_Notify,text="Clear Filters")
                    Clear_Filterbutton.place(x=360, y=95)
                   

                    self.export_Table['columns'] = (
                    "Invoice Number","Purchase ID", "Item", "Quantity", "Total Price", "Discount", "Date Purchased")
                    self.export_Table.column("#0", width=0, stretch=NO)
                    self.export_Table.column("Invoice Number", anchor=W,width=60)
                    self.export_Table.column("Purchase ID", anchor=W,width=50)
                    self.export_Table.column("Item", anchor=W,width=330)
                    self.export_Table.column("Quantity", anchor=E,width=30)
                    self.export_Table.column("Total Price", anchor=E,width=30)
                    self.export_Table.column("Discount", anchor=E,width=30)
                    self.export_Table.column("Date Purchased", anchor=E,width=50)

                    self.export_Table.heading("#0")
                    self.export_Table.heading("Invoice Number", text="Invoice Number", anchor=W)
                    self.export_Table.heading("Purchase ID", text="Purchase ID", anchor=W)
                    self.export_Table.heading("Item", text="Item", anchor=W)
                    self.export_Table.heading("Quantity", text="Quantity", anchor=W)
                    self.export_Table.heading("Total Price", text="Total Price", anchor=W)
                    self.export_Table.heading("Discount", text="Discount", anchor=W)
                    self.export_Table.heading("Date Purchased", text="Date Purchased", anchor=W)

                    self.export_Table.pack()

                elif report_type == 'Inventory':
                    try:
                        Filterbutton.place_forget()
                        Clear_Filterbutton.place_forget()
                        Deli_Filterbutton.place_forget()
                        Deli_Clear_Filterbutton.place_forget()
                    except(NameError):
                        pass

                    Inven_Filterbutton=Button(self.Add_Notify,text="Inventory Filter",command=self.Inven_Filter_GUI)
                    Inven_Filterbutton.place(x=200, y=95)
                    Inven_Clear_Filterbutton=Button(self.Add_Notify,text="Clear Filters")
                    Inven_Clear_Filterbutton.place(x=360, y=95)
        
                    self.export_Table['columns'] = (
                    'Reference ID', "Item", "Price", "Remaining Quantity")
                    self.export_Table.column("#0", width=0, stretch=NO)
                    self.export_Table.column("Reference ID", anchor=W,width=50)
                    self.export_Table.column("Item", anchor=W,width=300)
                    self.export_Table.column("Price", anchor=E,width=50)
                    self.export_Table.column("Remaining Quantity", anchor=E,width=70)
                        
                    self.export_Table.heading("#0")
                    self.export_Table.heading("Reference ID", text="Reference ID", anchor=W)
                    self.export_Table.heading("Item", text="Item", anchor=W)
                    self.export_Table.heading("Price", text="Price", anchor=W)
                    self.export_Table.heading("Remaining Quantity", text="Remaining Quantity", anchor=W)

                    self.export_Table.pack()

                elif report_type == "Delivery":
                    try:
                        Filterbutton.place_forget()
                        Clear_Filterbutton.place_forget()
                        Inven_Filterbutton.place_forget()
                        Inven_Clear_Filterbutton.place_forget()
                    except(NameError):
                        pass

                    Deli_Filterbutton=Button(self.Add_Notify,text="Delivery Filter",command=self.Del_Filter_GUI)
                    Deli_Filterbutton.place(x=200, y=95)
                    Deli_Clear_Filterbutton=Button(self.Add_Notify,text="Clear Filters")
                    Deli_Clear_Filterbutton.place(x=360, y=95)

                    self.export_Table['columns'] = (
                    'Batch Code', 'Item', 'Quantity', 'Price', 'Status')
                    self.export_Table.column("#0", width=0, stretch=NO)
                    self.export_Table.column("Batch Code", anchor=W,width=50 )
                    self.export_Table.column("Item", anchor=W,width=300 )
                    self.export_Table.column("Quantity", anchor=E,width=30 )
                    self.export_Table.column("Price", anchor=E,width=30 )
                    self.export_Table.column("Status", anchor=E,width=50 )
                        
                    self.export_Table.heading("#0")
                    self.export_Table.heading("Batch Code", text="Batch Code", anchor=W)
                    self.export_Table.heading("Item", text="Item", anchor=W)
                    self.export_Table.heading("Quantity", text="Quantity", anchor=W)
                    self.export_Table.heading("Price", text="Price", anchor=W)
                    self.export_Table.heading("Status", text="Status", anchor=W)

                    self.export_Table.pack()

                # if report_type == 'Forecast':
                #     try:
                #         FLabel.place_forget()
                #         TOLabel.place_forget()
                #         YLabel.place_forget()
                #         scope.place_forget()
                #         scope_to.place_forget()
                #         scope_year.place_forget()
                #         Radio_Day.place_forget()
                #         Radio_Monthly.place_forget()
                #         Radio_TO.place_forget()
                #         Radio_Yearly.place_forget()
                #     except(NameError):
                #         pass

                #     self.export_Table['columns'] = (
                #     'Id','Item','Quantity','Price','30 Day Forecast')
                #     self.export_Table.column("#0", width=0, stretch=NO)
                #     self.export_Table.column("Id", anchor=W, width=50)
                #     self.export_Table.column("Item", anchor=W,width=340 )
                #     self.export_Table.column("Quantity", anchor=E,width=30 )
                #     self.export_Table.column("Price", anchor=E,width=30 )
                #     self.export_Table.column("30 Day Forecast", anchor=E,width=50 )

                #     self.export_Table.heading("#0")
                #     self.export_Table.heading("Id", text="Id", anchor=W)
                #     self.export_Table.heading("Item", text="Item", anchor=W)
                #     self.export_Table.heading("Quantity", text="Quantity", anchor=W)
                #     self.export_Table.heading("Price", text="Price", anchor=W)
                #     self.export_Table.heading("30 Day Forecast", text="30 Day Forecast", anchor=W)

                #     self.export_Table.pack()

                count = 0
                for item in result:
                        self.export_Table.insert('', index='end', iid=count, text=item, values=(item))
                        count += 1
                
                    # if report_type == 'Forecast':
                    #     for item in range(len(result)):
                    #         self.export_Table.insert('', index='end', iid=count, text=(item,values), values=(result[item][0],result[item][1],result[item][2],result[item][3],values))
                    #         count += 1
                    # else:
                    #     for item in result:
                    #         self.export_Table.insert('', index='end', iid=count, text=item, values=(item))
                    #         count += 1

            reports.bind('<<ComboboxSelected>>',reports_callback)
            PageOpen += 1
            self.Add_Notify.mainloop()
        else:
            messagebox.showinfo("Error","The Window already Open!")

    # Chick ADD END!~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    def InvorGUI(self):
        self.InvorVal = Tk()
        self.InvorVal.title("Cresdel Pharmacy!!")
        width = self.InvorVal.winfo_screenwidth()
        height = self.InvorVal.winfo_screenheight()
        self.InvorVal.geometry("%dx%d" % (width, height))
        # self.InvorVal.resizable(False, False)

        # For the Page 1 Detail
        self.Frame_Detail = Frame(self.InvorVal, width=1063,bg="yellow", height=200, highlightbackground="black",
                                  highlightthickness=1, padx=10, pady=10)
        self.Frame_Detail.place(x=0, y=0)

        Title_label=Label(self.Frame_Detail,text="Cresdel Pharmacy!!",bg="yellow",font=("Arial", 50, "bold"))
        Title_label.place(x=20,y=50)

        # For the Page LIST
        self.Frame_main = Frame(self.InvorVal, width=1063, height=540, highlightbackground="black",
                                highlightthickness=2)
        self.Frame_main.place(x=0, y=199)
        self.Frame_List = Frame(self.Frame_main, width=1058, height=540, padx=10, pady=10)
        self.Frame_stack = Frame(self.Frame_main, width=1058, height=540, padx=10, pady=10)
        self.Frame_Empl = Frame(self.Frame_main, width=1058, height=540, padx=10, pady=10)
        self.Frame_Del = Frame(self.Frame_main, width=1058, height=540, padx=10, pady=10)

        # For the Side
        self.Frame_Side = Frame(self.InvorVal, width=300, height=743, highlightbackground="black", highlightthickness=3,
                                padx=10, pady=10)
        self.Frame_Side.place(x=1060, y=0)

        Detail = Label(self.Frame_Side, text="Detail Button", width=37, anchor=W).place(x=40, y=270)
        self.button_List = Button(self.Frame_Side, text="List", padx=10, pady=10, width=10, height=1, bg='#54FA9B',
                                  command=self.Click_List1)
        self.button_Stack = Button(self.Frame_Side, text="Sales", padx=10, pady=10, width=10, height=1, bg='#54FA9B',
                                   command=self.Click_Stack)
        self.button_Delivery = Button(self.Frame_Side, text="Delivery", padx=10, pady=10, width=10, height=1,
                                      bg='#54FA9B', command=self.Click_Delivery)
        self.button_Employee = Button(self.Frame_Side, text="Employee", padx=10, pady=10, width=10, height=1,
                                      bg='#54FA9B', command=self.Click_Employee)

        label = Label(self.Frame_Side, text="IMAGE").place(x=100, y=10)
        Add = Label(self.Frame_Side, text="ADD Button", width=37, anchor=W).place(x=40, y=420)

        self.Add_Del = Button(self.Frame_Side, text="ADD Delivery", padx=10, pady=10, width=10, height=1, bg='#54FA9B',
                              command=self.Click_Add_Delivery)
        self.button_Add_Em = Button(self.Frame_Side, text="ADD Employee", padx=10, pady=10, width=10, height=1,
                                    bg='#54FA9B', command=self.Click_Add_Em)


        self.button_Add_Pm = Button(self.Frame_Side, text="ADD Product", padx=10, pady=10, width=10, height=1,
                                    bg='#54FA9B', command=self.Click_Add_Product)
        self.btn_Notification = Button(self.Frame_Side, text="Export", padx=10, pady=10, width=10, height=1,
                                       bg='#54FA9B', command=self.notify_UI)
        button_Out= Button(self.Frame_Side,text="Back Home",padx=10,pady=10,width=10,height=1,bg='#54FA9B',command=self.InvorVal.destroy)
        button_Out.place(x=160,y=600)

        def goto_forecast():
            forecast.GUI()

        button_Out= Button(self.Frame_Side,text="Forecast",padx=10,pady=10,width=10,height=1,bg='#54FA9B',command=goto_forecast)
        button_Out.place(x=40,y=600)

        self.button_List.place(x=40, y=300)
        self.button_Stack.place(x=160, y=300)
        self.button_Delivery.place(x=40, y=350)
        self.button_Employee.place(x=160, y=350)

        self.Add_Del.place(x=40, y=450)
        self.button_Add_Em.place(x=160, y=450)
        self.button_Add_Pm.place(x=40, y=500)
        # self.button_Add_prodref.place(x=160,y=500)
        self.btn_Notification.place(x=160, y=500)
    
        self.Click_List1()

        self.InvorVal.mainloop()
   
    def Edit_Stack_close(self):
            global PageOpen_Sub
            self.Edit_Stack.wm_attributes("-topmost", 0)
            if messagebox.askokcancel('Close', 'Are you sure you want to close the Add Product Page all the data will not be Save?'):
                self.Add_Stack.wm_attributes("-topmost", 1)
                self.Add_Stack.grab_release()
                PageOpen_Sub=1
                self.Edit_Stack.destroy()
            else:
                self.Edit_Stack.wm_attributes("-topmost", 1)
                
    def Click_Edit_Ref(self, var):
        if var == "Select Product":
            self.Add_Stack.wm_attributes("-topmost", 0)
            messagebox.showerror("An Error Occured","Pick an Item by Searching it")
            self.Add_Stack.wm_attributes("-topmost", 1)
        else:
            global PageOpen_Sub
            self.Add_Stack.wm_attributes("-topmost", 0)
            if PageOpen_Sub<2:

                self.Edit_Stack = Toplevel(self.Add_Stack)
                self.Edit_Stack.title("Edit Product Reference")
                self.Edit_Stack.geometry("700x350")
                self.Edit_Stack.protocol("WM_DELETE_WINDOW",self.Edit_Stack_close)
                self.Edit_Stack.wm_attributes("-topmost", 1)
                self.Edit_Stack.resizable(False, False)
                self.Edit_Stack.grab_set()


                self.Edit_Frame_Product = Frame(self.Edit_Stack, width=700, height=350, )
                self.Edit_Frame_Product.grid(row=0, column=0)

                self.Frma = Label(self.Edit_Frame_Product, text="Edit Product", width=20, font=("Arial", 35), anchor=W)
                self.Frma.place(x=20, y=10)

                a = Product.product()
                lst = a.return_one(var)
                idd = lst[0][0]
                namee = lst[0][1]
                pricee = lst[0][2]

                self.Old = Label(self.Edit_Frame_Product, text="Old Product Name", width=20, font=("Arial", 15), anchor=W)
                self.Old.place(x=50, y=70)

                global ref_id_entry, name, price
                ref_id_entry = StringVar()
                name = StringVar()
                price = StringVar()

                self.Stack_Product_ID_Label = Label(self.Edit_Frame_Product, text="ID:").place(x=50, y=110)
                self.Stack_Product_ID_EN = Entry(self.Edit_Frame_Product, width=10, borderwidth=5, textvariable=ref_id_entry,
                                                state="disabled")
                self.Stack_Product_ID_EN.place(x=50, y=130)

                self.Stack_Product_Item_Label = Label(self.Edit_Frame_Product, text="Product Name:", ).place(x=130, y=110)
                self.Stack_Product_Item_ENN = Entry(self.Edit_Frame_Product, width=70, textvariable=name, borderwidth=5,
                                                    state="disabled")
                self.Stack_Product_Item_ENN.place(x=130, y=130)

                self.Stack_Product_ID_Label = Label(self.Edit_Frame_Product, text="Price:").place(x=570, y=110)
                self.Stack_Product_PRICE_EN = Entry(self.Edit_Frame_Product, textvariable=price, width=15, borderwidth=5,
                                                    state="disabled")
                self.Stack_Product_PRICE_EN.place(x=570, y=130)

                self.Stack_Product_ID_EN.config(state='normal')
                self.Stack_Product_Item_ENN.config(state='normal')
                self.Stack_Product_PRICE_EN.config(state='normal')

                ref_id_entry.set(idd)
                self.Stack_Product_ID_EN.insert(0,ref_id_entry.get())
                name.set(namee)
                self.Stack_Product_Item_ENN.insert(0,name.get())
                price.set(pricee)
                self.Stack_Product_PRICE_EN.insert(0,price.get())

                self.Stack_Product_ID_EN.config(state='disabled')
                self.Stack_Product_Item_ENN.config(state='disabled')
                self.Stack_Product_PRICE_EN.config(state='disabled')

                self.New = Label(self.Edit_Frame_Product, text="New Product Name", width=20, font=("Arial", 15), anchor=W)
                self.New.place(x=50, y=180)

                global ref_name_entry
                ref_name_entry = StringVar()
                self.Stack_Product_Name_Label = Label(self.Edit_Frame_Product, text="New Product Name:").place(x=50, y=210)
                self.Stack_Product_Name_ENN = Entry(self.Edit_Frame_Product, width=83, borderwidth=5,
                                                    textvariable=ref_name_entry)
                self.Stack_Product_Name_ENN.place(x=50, y=230)

                global ref_price_entry
                ref_price_entry = StringVar()
                self.Stack_Product_Name_Label = Label(self.Edit_Frame_Product, text="New Product Price:").place(x=570, y=210)
                self.Stack_Product_Price_EN = Entry(self.Edit_Frame_Product, width=15, borderwidth=5,
                                                    textvariable=ref_price_entry)
                self.Stack_Product_Price_EN.place(x=570, y=230)

                self.submit = Button(self.Edit_Frame_Product, text="Submit Changes", padx=20, pady=5,
                                    command=self.Click_ref_submit).place(x=450, y=280)

                self.button_Out = Button(self.Edit_Frame_Product, text="Cancel", padx=9, pady=5, bg="green",
                                        command=self.Edit_Stack_close)
                self.button_Out.place(x=600, y=280)
                PageOpen_Sub += 1
            else:
                messagebox.showinfo("Error","The Window is already Open!")

    def Click_ref_submit(self):

        id = ref_id_entry.get()
        name = self.Stack_Product_Name_ENN.get()
        price = self.Stack_Product_Price_EN.get()

        idd = int(id)
        pricee = float(price)
        priceee = float(pricee)

        Prod = Product.product()
        Prod.editReference(idd, name, priceee)
        self.Edit_Stack.wm_attributes("-topmost", 0)
        messagebox.showinfo("Edit Success","Edit Success!")
        self.Edit_Stack.wm_attributes("-topmost", 1)

    def setRefVals(self, event):
        global choice
        choice = event.widget.get()
        res = [idx for idx, x in enumerate(lst) if x[1] == choice]
        item = lst[res[0]]
        price = item[2]
        name = item[1]
        id = item[0]

        ref_price_entry.set(price)
        ref_name_entry.set(name)
        ref_id_entry.set(id)

    def Add_Ref_close(self):
            global PageOpen_Sub
            self.Add_Stack_ADD.wm_attributes("-topmost", 0)
            if messagebox.askokcancel('Close', 'Are you sure you want to close the Add Product Page all the data will not be Save?'):
                self.Add_Stack.wm_attributes("-topmost", 1)
                self.Add_Stack.grab_release()
                PageOpen_Sub=1
                self.Add_Stack_ADD.destroy()
            else:
                self.Add_Stack_ADD.wm_attributes("-topmost", 1)

    
    def Click_Add_Ref(self):
        global PageOpen_Sub
        self.Add_Stack.wm_attributes("-topmost", 0)
        if PageOpen_Sub<2:
            self.Add_Stack_ADD = Toplevel(self.Add_Stack)
            self.Add_Stack_ADD.title("Add Product Reference")
            self.Add_Stack_ADD.geometry("700x543")
            self.Add_Stack_ADD.resizable(False,False)
            self.Add_Stack_ADD.protocol("WM_DELETE_WINDOW",self.Add_Ref_close)
            self.Add_Stack_ADD.wm_attributes("-topmost", 1)
            self.Add_Stack_ADD.grab_set()

            self.Frame_Add_St = Frame(self.Add_Stack_ADD, width=700, height=350, )
            self.Frame_Add_St.grid(row=0, column=0)

            self.Frma = Label(self.Frame_Add_St, text="Add Product!!", width=20, font=("Arial", 40), anchor=W)
            self.Frma.place(x=20, y=10)

            self.Stack_Product_Name_Label = Label(self.Frame_Add_St, text="Product Name:").place(x=30, y=90)
            self.Stack_Product_Name_EN = Entry(self.Frame_Add_St, width=70, borderwidth=5)
            self.Stack_Product_Name_EN.place(x=30, y=110)

            global price_entry
            price_entry = tk.StringVar()
            self.Stack_Product_Price_Label = Label(self.Frame_Add_St, text="Price:").place(x=470, y=90)
            self.Stack_Product_Price_EN = Entry(self.Frame_Add_St, width=15, borderwidth=5, textvariable=price_entry)
            self.Stack_Product_Price_EN.place(x=470, y=110)

            self.Frame_ListS = Frame(self.Add_Stack_ADD, width=800, height=320, highlightbackground="black",
                                    highlightthickness=3, padx=5, pady=5)
            self.Frame_ListS.place(x=0, y=200)
            # Table
            self.frame_Table = ttk.Treeview(self.Frame_ListS, height=15)
            self.frame_Table['columns'] = ("ID", "Name", "Price")
            self.frame_Table.column("#0", width=0, stretch=NO)
            self.frame_Table.column("ID", anchor=W, width=100)
            self.frame_Table.column("Name", anchor=W, width=482)
            self.frame_Table.column("Price", anchor=E, width=100)
            # Table Head
            self.frame_Table.heading("#0")
            self.frame_Table.heading("ID", text="ID", anchor=W)
            self.frame_Table.heading("Name", text="Product Name", anchor=W)
            self.frame_Table.heading("Price", text="Price", anchor=W)

            self.frame_Table.pack(fill='both')
            self.frame_Table.grid(row=1, column=0)

            self.button_Add = Button(self.Frame_Add_St, text="Add", padx=20, pady=5, command=self.reference_Done)
            self.button_Add.place(x=360, y=150)

            self.button_Delete = Button(self.Frame_Add_St, text="Delete", padx=20, pady=5, command=self.Delete)
            self.button_Delete.place(x=440, y=150)

            self.button_Finish = Button(self.Frame_Add_St, text="Finish", padx=20, pady=5, state='disabled')
            self.button_Finish.place(x=530, y=150)

            self.button_Cancel = Button(self.Frame_Add_St, text="Cancel", padx=9, pady=5, bg="green",
                                        command=self.Add_Ref_close)
            self.button_Cancel.place(x=620, y=150)
            PageOpen_Sub+=1
        else:
            messagebox.showinfo("Error","The Window is already Open!")

    def reference_Done(self):
        ProductName = self.Stack_Product_Name_EN.get()
        price = self.Stack_Product_Price_EN.get()
        id = randomNumGen.generateProductID()

        pricee = int(price)
        priceee = float(pricee)

        global vals
        vals = (id, ProductName, priceee)

        global count

        id = []
        name = []
        price = []

        if 'count' not in globals():
            count = 0
        else:
            count += 1
        self.frame_Table.insert(parent='', index='end', iid=count, text=vals, values=(vals))
        vals = ()

        for child in self.frame_Table.get_children():
            val = self.frame_Table.item(child)["values"]
            idd = randomNumGen.generateProductID()

            id.append(idd)
            name.append(val[1])
            price.append(val[2])

            vals = list(zip(id, name, price))

        self.button_Finish.config(state='normal', command=self.AddReference)

    def AddReference(self):
        self.frame_Table.delete(*self.frame_Table.get_children())
        Prod = Product.product()
        Prod.addReference(vals)

    def Click_List(self):
        global Entry_Search
        global Search_Table
        window_list = Toplevel()
        window_list.title("Delivery Batch")
        window_list.geometry("500x430")

        window_Frame = Frame(window_list, width=400, height=100)
        window_Frame.grid(row=0, column=0)

        window_Frame2 = Frame(window_list, width=400, height=250, bg="blue")
        window_Frame2.grid(row=1, column=0)

        Search_Table = ttk.Treeview(window_Frame2, height=12)
        Search_Table['column'] = ("Batch Code", "Order Date", "Arrival Date")
        Search_Table.column("#0", width=0, stretch=NO, anchor=W)
        Search_Table.column("Batch Code", width=100, stretch=NO, anchor=W)
        Search_Table.column("Order Date", width=148, stretch=NO, anchor=W)
        Search_Table.column("Arrival Date", width=100, stretch=NO, anchor=E)

        Search_Table.heading("#0")
        Search_Table.heading("Batch Code", text="Batch Code", anchor=W)
        Search_Table.heading("Order Date", text="Order Date", anchor=W)
        Search_Table.heading("Arrival Date", text="Arrival Date", anchor=W)

        Search_Table.grid(row=0, column=0)

        Label_Search = Label(window_Frame, text="Search Batch:")
        Entry_Search = Entry(window_Frame, width=50, borderwidth=3)
        button_Search = Button(window_Frame, text="Mark On Hand", padx=5, pady=0, command=self.search)

        Label_Search.grid(row=0, column=0, sticky=W)
        Entry_Search.grid(row=0, column=1)
        button_Search.grid(row=0, column=3)

        m = Employee.Employee()
        m1 = m.ListAllBatches()
        count = 0

        for x in m1:
            count += 1
            Search_Table.insert(parent='', index='end', iid=count, text=x, values=x)

    def close_window(self):
        self.Add_Del.config(state='normal')
        self.button_Add_Em.config(state='normal')
        self.button_Add_Pm.config(state='normal')
        # self.btn_Notification.config(state='normal')
        frame.destroy()

    def start(self, id, role):
        global user_role, user_id
        user_role=role
        user_id=id
        self.InvorGUI()

